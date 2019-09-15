'''==========================================
Title:  folder_wizard.py
Author:  Bryan Bordeman
Start Date:  062219
Updated:  071019
Version:  v2.0

;=========================================='''

import pickle
import os
import time
from get_next_num import get_next_quote_num
from get_next_num import get_next_project_num
from openpyxl import*
from zip_2_state import find_state
from project_attribute_list import*

# global var ----------------------------
current_year = time.strftime("%Y")

# below path is for testing at home
project_dir = r'C:\Users\Bryan\Google Drive\Programming\Python\folder_wizard\Global'
opportunity_dir = r"C:\Users\Bryan\Google Drive\Programming\Python\folder_wizard\RFQ's"

# below path is live on office server
# project_dir = r'T:\Global'
# opportunity_dir = r"T:\RFQ's"


def main():
    pass
# below is for testing only
#     quote_data = 'opportunity.pkl'
#     quote_obj = {}  # if quote_data does not exist
#     if os.path.exists(quote_data):
#         with open(quote_data, 'rb') as rfp:
#             quote_obj = pickle.load(rfp)
#     print(quote_obj)


class Opportunity(object):
    def __init__(self, project_name, project_category, project_type, type_code, project_zip, customer_list, bid_due, manager):
        self.quote_number = get_next_quote_num()
        self.project_name = project_name[:27]
        self.project_category = project_category
        self.project_type = project_type
        self.type_code = type_code
        self.project_zip = project_zip
        self.customer_list = customer_list
        self.bid_due = bid_due
        self.manager = manager

        quote = f'{self.quote_number} {self.manager} {self.project_name} {self.type_code}'

        self.write_log()
        self.pickle_opp()

        # make new year folder if current year does not match dir list-------------
        opportunity_dir_list = (os.listdir(opportunity_dir))
        year_list = []

        for year in opportunity_dir_list:
            try:
                if int(year[:5]):
                    year_list.append(year)
            except ValueError:
                continue
        year_list.sort()
        if year_list[-1] == str(current_year):
            year_dir = current_year
        else:
            self.createFolder(f'{opportunity_dir}/{current_year} Quotes')

        #-----------------------------------------------------------------------

        self.createFolder(f'{opportunity_dir}/{current_year} Quotes/{quote}/00_quotations_estimates')
        self.createFolder(f'{opportunity_dir}/{current_year} Quotes/{quote}/01_drawings_specs')
        self.createFolder(f'{opportunity_dir}/{current_year} Quotes/{quote}/02_rfi_addenda')
        self.createFolder(f'{opportunity_dir}/{current_year} Quotes/{quote}/03_photos')
        self.createFolder(f'{opportunity_dir}/{current_year} Quotes/{quote}/04_misc_docs')
        self.make_readme(quote)

    def __repr__(self):
        return f'{self.quote_number} {self.manager} {self.project_name} {self.type_code}'

    def pickle_opp(self):
        '''Save data in dictionary so it easy to retrieve later.'''
        quote_data = 'opportunity.pkl'
        quote_obj = {}  # if quote_data does not exist
        if os.path.exists(quote_data):
            with open(quote_data, 'rb') as rfp:
                quote_obj = pickle.load(rfp)

        # Look up will be as follows:
        # quote_obj[quote_number][index]
        # index 0 = project_name
        # index 1 = project_category
        # index 2 = project_type
        # index 3 = type_code
        # index 4 = project_zip
        # index 5 = customer_list
        # index 6 = bid_due]

        quote_obj[self.quote_number] = [self.project_name, self.project_category, self.project_type, self.type_code, self.project_zip, self.customer_list, self.bid_due]

        with open(quote_data, 'wb') as wfp:
            pickle.dump(quote_obj, wfp)

    def write_log(self):
        '''update quote log'''
        log = "Master Quote Log"
        completeName = os.path.join(opportunity_dir, log + ".xlsx")
        book = load_workbook(completeName)
        ws = book.worksheets[0]
        for cell in ws["A"]:
            if cell.value is None:
                current_row = cell.row
                break
        else:
            cell.row + 1

        ws[f'A{current_row}'] = self.quote_number
        ws[f'B{current_row}'] = self.manager
        ws[f'C{current_row}'] = time.strftime("%D")
        ws[f'D{current_row}'] = f'{self.project_name} {self.type_code}'
        ws[f'E{current_row}'] = find_state(int(self.project_zip))
        ws[f'F{current_row}'] = self.bid_due

        book.save(completeName)
        book.close()

    def createFolder(self, directory):
        try:
            if not os.path.exists(directory):
                os.makedirs(directory)
        except OSError:
            print('Error: Creating directory. ' + directory)

    def make_readme(self, quote):
        name_of_file = 'README'
        path = f'{opportunity_dir}/{current_year} Quotes/{quote}'
        completeName = os.path.join(path, name_of_file + ".txt")

        readme = open(completeName, "w")

        opportunity_info = f'Quote Number = {self.quote_number}\nProject Name = {self.project_name}\nProject Category = {self.project_category}\nProject Type = {self.project_type}\nProject Zip = {self.project_zip}\nBid Due Date = {self.bid_due}\nCustomer List = {self.customer_list}'

        # f'Project Number = {self.project_number}\nProject Name = {self.project_name}\nProject Category = {self.project_category}\nProject Type = {self.project_type}\nProject Zip = {self.project_zip}\nCustomer = {self.customer}\nQuote Number = {self.quote}\nTerms = {self.terms}\nTax Exempt = {self.tax}\nBilling Type = {self.billing}\nSell Price (USD) = ${self.price}\n'
        readme.write(opportunity_info)
        readme.close()


class Project(object):
    def __init__(self, project_name, project_category, project_type, type_code, project_zip, customer, quote, terms, tax, billing, labor_code, order_type, price):
        self.project_number = get_next_project_num()
        self.project_name = project_name[:27]
        self.project_category = project_category
        self.project_type = project_type
        self.type_code = type_code
        self.project_zip = project_zip
        self.customer = customer
        self.quote = quote
        self.terms = terms
        self.tax = tax  # set as  boolean
        self.billing = billing
        self.labor_code = labor_code
        self.order_type = order_type
        self.price = price

        project = f'{self.project_number} {self.project_name} {self.type_code}'

        self.write_log()
        self.pickle_project()

        # make new year folder if current year does not match dir list-------------
        project_dir_list = (os.listdir(project_dir))
        year_list = []

        for year in project_dir_list:
            try:
                if int(year[:5]) and len(year) < 5:
                    year_list.append(year)
            except ValueError:
                continue
        year_list.sort()
        if year_list[-1] == str(current_year):
            year_dir = current_year
        else:
            self.createFolder(f'{project_dir}/{current_year}')
        #--------------------------------------------------------------------------

        self.createFolder(f'{project_dir}/{current_year}/{project}/photos')
        self.createFolder(f'{project_dir}/{current_year}/{project}/test_reports')
        self.createFolder(f'{project_dir}/{current_year}/{project}/insurance_docs')
        self.createFolder(f'{project_dir}/{current_year}/{project}/minutes_etc')
        self.createFolder(f'{project_dir}/{current_year}/{project}/travel')
        self.createFolder(f'{project_dir}/{current_year}/{project}/tx')
        self.createFolder(f'{project_dir}/{current_year}/{project}/billing')
        self.createFolder(f'{project_dir}/{current_year}/{project}/production')
        self.createFolder(f'{project_dir}/{current_year}/{project}/RFIs')
        self.createFolder(f'{project_dir}/{current_year}/{project}/Purchasing')
        self.createFolder(f'{project_dir}/{current_year}/{project}/Material_Specs')
        self.createFolder(f'{project_dir}/{current_year}/{project}/contracts')
        self.createFolder(f'{project_dir}/{current_year}/{project}/contracts/change_orders')
        self.createFolder(f'{project_dir}/{current_year}/{project}/contracts/closeout_documents')
        self.createFolder(f'{project_dir}/{current_year}/{project}/contracts/AIA_docs_for_pay_apps')
        self.createFolder(f'{project_dir}/{current_year}/{project}/contracts/backup_and_old_files')
        self.createFolder(f'{project_dir}/{current_year}/{project}/contracts/Exhibits')
        self.createFolder(f'{project_dir}/{current_year}/{project}/contracts/QB_Invoices')
        self.createFolder(f'{project_dir}/{current_year}/{project}/contracts/TX_Ex')
        self.createFolder(f'{project_dir}/{current_year}/{project}/drawings')
        self.createFolder(f'{project_dir}/{current_year}/{project}/drawings/drawings_sent')
        self.createFolder(f'{project_dir}/{current_year}/{project}/drawings/revisions')
        self.createFolder(f'{project_dir}/{current_year}/{project}/drawings/archive_dwgs')
        self.createFolder(f'{project_dir}/{current_year}/{project}/drawings/approved_dwgs')
        self.createFolder(f'{project_dir}/{current_year}/{project}/drawings/arch_dwgs')

        self.make_readme(project)

    def pickle_project(self):
        '''Save data in dictionary so it easy to retrieve later.'''
        project_data = 'project.pkl'
        project_obj = {}  # if quote_data does not exist
        if os.path.exists(project_data):
            with open(project_data, 'rb') as rfp:
                project_obj = pickle.load(rfp)

        # Look up will be as follows:
        # project_obj[project_number][index]
        # index 0 = project_name
        # index 1  = project_category
        # index 2  = project_type
        # index 3  = type_code
        # index 4  = project_zip
        # index 5  = customer
        # index 6  = quote
        # index 7  = terms
        # index 8  = tax  # set as  boolean
        # index 9  = billing
        # index 10  = labor_code

        project_obj[self.project_number] = [self.project_name, self.project_category, self.project_type, self.type_code, self.project_zip, self.customer, self.quote, self.terms, self.tax, self.billing, self.labor_code]

        with open(project_data, 'wb') as wfp:
            pickle.dump(project_obj, wfp)

    def __repr__(self):
        return f'{self.project_number} {self.project_name} {self.type_code}'

    def write_log(self):

        log = "Global Job List-start-complete dates 2016"

        completeName = os.path.join(project_dir, log + ".xlsx")

        book = load_workbook(completeName)
        ws = book.worksheets[0]
        for cell in ws["B"]:
            if cell.value is None:
                current_row = cell.row
                break
        else:
            cell.row + 1

        ws[f'A{current_row}'] = f'{self.project_number}{self.billing}'
        ws[f'B{current_row}'] = find_state(int(self.project_zip)).split(', ')[-1]
        ws[f'C{current_row}'] = self.order_type
        ws[f'D{current_row}'] = f'{self.project_name} {self.type_code}'
        ws[f'E{current_row}'] = self.labor_code
        ws[f'F{current_row}'] = time.strftime("%D")
        ws[f'G{current_row}'] = self.project_type

        book.save(completeName)
        book.close()

    def createFolder(self, directory):
        try:
            if not os.path.exists(directory):
                os.makedirs(directory)
        except OSError:
            print('Error: Creating directory. ' + directory)

    def make_readme(self, project):
        name_of_file = 'README'
        path = f'{project_dir}/{current_year}/{project}'
        completeName = os.path.join(path, name_of_file + ".txt")

        readme = open(completeName, "w")

        # convert boolean to yes/ no string for readme
        if self.tax == 1:
            self.tax = 'Yes'
        elif self.tax == 0:
            self.tax = 'No'

        # convert billing back to key for readme
        for key, value in billing_dict.items():
            if value == self.billing:
                self.billing = key

        project_info = f'Project Number = {self.project_number}\nProject Name = {self.project_name}\nProject Category = {self.project_category}\nProject Type = {self.project_type}\nProject Zip = {self.project_zip}\nCustomer = {self.customer}\nQuote Number = {self.quote}\nTerms = {self.terms}\nTax Exempt = {self.tax}\nBilling Type = {self.billing}\nSell Price (USD) = ${self.price}\n'
        readme.write(project_info)
        readme.close()


if __name__ == "__main__":
    main()
